from __future__ import annotations

import argparse
import math
from pathlib import Path

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.colors import LinearSegmentedColormap, TwoSlopeNorm
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "outputs" / "pilot" / "joh_hardening_v2"
OUT = DATA / "production_figures_v2"

EVENTS = ["Q95", "Q98", "Q99"]
HORIZONS = [1, 5, 7]
SOURCES = ["AUS", "CH", "COL", "DE", "DK", "ES", "FR", "IND", "KR", "NZ/FI"]

COLORS = {
    "anchor": "#577590",
    "dala": "#008F83",
    "current": "#C93C3C",
    "near": "#E9C46A",
    "above": "#2A9D8F",
    "far": "#8D99AE",
    "ink": "#263238",
    "grid": "#D9E1E5",
}


def setup_style() -> None:
    mpl.rcParams.update(
        {
            "font.family": "DejaVu Sans",
            "font.size": 9.5,
            "axes.titlesize": 10.5,
            "axes.labelsize": 9.5,
            "xtick.labelsize": 8.5,
            "ytick.labelsize": 8.5,
            "axes.edgecolor": COLORS["ink"],
            "axes.labelcolor": COLORS["ink"],
            "xtick.color": COLORS["ink"],
            "ytick.color": COLORS["ink"],
            "text.color": COLORS["ink"],
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
            "savefig.facecolor": "white",
        }
    )


def panel(ax: plt.Axes, label: str, title: str) -> None:
    ax.text(-0.13, 1.08, label, transform=ax.transAxes, fontsize=12, fontweight="bold", va="top")
    ax.set_title(title, loc="left", pad=8, fontweight="bold")


def clean_axes(ax: plt.Axes, grid_axis: str | None = "y") -> None:
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    if grid_axis:
        ax.grid(axis=grid_axis, color=COLORS["grid"], linewidth=0.7, zorder=0)
        ax.set_axisbelow(True)


def heatmap(
    ax: plt.Axes,
    matrix: np.ndarray,
    row_labels: list[str],
    col_labels: list[str],
    *,
    fmt: str = "+.3f",
    cmap: str | mpl.colors.Colormap = "YlGnBu",
    center: float | None = None,
    vmin: float | None = None,
    vmax: float | None = None,
    text: np.ndarray | None = None,
    cbar_label: str = "",
) -> mpl.image.AxesImage:
    finite = matrix[np.isfinite(matrix)]
    if finite.size == 0:
        raise RuntimeError("Heatmap contains no finite values")
    if center is not None:
        lo = float(np.nanmin(matrix)) if vmin is None else vmin
        hi = float(np.nanmax(matrix)) if vmax is None else vmax
        if lo >= center:
            lo = center - max(abs(hi - center) * 0.08, 1e-6)
        if hi <= center:
            hi = center + max(abs(center - lo) * 0.08, 1e-6)
        norm = TwoSlopeNorm(vmin=lo, vcenter=center, vmax=hi)
        image = ax.imshow(matrix, cmap=cmap, norm=norm, aspect="auto")
    else:
        image = ax.imshow(matrix, cmap=cmap, vmin=vmin, vmax=vmax, aspect="auto")
    ax.set_xticks(np.arange(len(col_labels)), col_labels)
    ax.set_yticks(np.arange(len(row_labels)), row_labels)
    ax.tick_params(length=0)
    for edge in ("top", "right", "bottom", "left"):
        ax.spines[edge].set_visible(False)
    values = text if text is not None else matrix
    threshold = (float(np.nanmin(matrix)) + float(np.nanmax(matrix))) / 2
    for r in range(matrix.shape[0]):
        for c in range(matrix.shape[1]):
            raw = matrix[r, c]
            shown = values[r, c]
            label = format(shown, fmt) if isinstance(shown, (int, float, np.number)) else str(shown)
            color = "white" if raw > threshold else COLORS["ink"]
            ax.text(c, r, label, ha="center", va="center", fontsize=8.5, color=color, fontweight="bold")
    cb = ax.figure.colorbar(image, ax=ax, fraction=0.046, pad=0.03)
    cb.outline.set_visible(False)
    cb.ax.tick_params(length=0, labelsize=8)
    if cbar_label:
        cb.set_label(cbar_label, fontsize=8.5)
    return image


def save_figure(fig: plt.Figure, number: int) -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    fig.savefig(OUT / f"Figure_{number}.pdf", bbox_inches="tight")
    fig.savefig(OUT / f"Figure_{number}.png", dpi=450, bbox_inches="tight")
    plt.close(fig)


def load_primary() -> pd.DataFrame:
    path = DATA / "candidate_unit_metrics_with_anchor_gains.csv"
    frame = pd.read_csv(path)
    frame = frame[
        frame["dataset_version"].eq("fully_deduplicated")
        & frame["candidate"].eq("op_lead_adaptive_direct")
    ].copy()
    expected = 10 * 5 * 3 * 3
    if len(frame) != expected or frame["source"].nunique() != 10:
        raise RuntimeError(f"Expected {expected} final DALA units across ten sources; found {len(frame)}")
    return frame


def source_seed_rollup(primary: pd.DataFrame) -> pd.DataFrame:
    return (
        primary.groupby(["event", "horizon", "seed"], sort=False)
        .agg(
            sources=("source", "nunique"),
            top2_hits_gain=("top2_daily_hits_gain_vs_anchor", "sum"),
            episode_gain=("top2_episodes_detected_gain_vs_anchor", "sum"),
        )
        .reset_index()
    )


def build_figure1(primary: pd.DataFrame) -> dict[str, pd.DataFrame]:
    cell = (
        primary.groupby(["event", "horizon"], sort=False)
        .agg(
            units=("daily_ap_gain_vs_anchor", "size"),
            median_ap_gain=("daily_ap_gain_vs_anchor", "median"),
            p10_ap_gain=("daily_ap_gain_vs_anchor", lambda x: float(x.quantile(0.10))),
            min_ap_gain=("daily_ap_gain_vs_anchor", "min"),
            nonnegative_units=("daily_ap_gain_vs_anchor", lambda x: int(x.ge(-1e-12).sum())),
        )
        .reset_index()
    )
    seed = source_seed_rollup(primary)
    q98_seed = seed[seed["event"].eq("Q98")].copy()
    q98_source = (
        primary[primary["event"].eq("Q98")]
        .groupby(["source", "horizon"], sort=False)
        .agg(
            seeds=("seed", "nunique"),
            median_top2_hit_gain=("top2_daily_hits_gain_vs_anchor", "median"),
            min_top2_hit_gain=("top2_daily_hits_gain_vs_anchor", "min"),
        )
        .reset_index()
    )

    ap = cell.pivot(index="event", columns="horizon", values="median_ap_gain").reindex(index=EVENTS, columns=HORIZONS)
    lower = cell.pivot(index="event", columns="horizon", values="p10_ap_gain").reindex(index=EVENTS, columns=HORIZONS)
    nn = cell.pivot(index="event", columns="horizon", values="nonnegative_units").reindex(index=EVENTS, columns=HORIZONS)
    source_matrix = q98_source.pivot(index="source", columns="horizon", values="median_top2_hit_gain").reindex(index=SOURCES, columns=HORIZONS)

    fig, axes = plt.subplots(2, 2, figsize=(11.4, 8.4), constrained_layout=True)
    ax = axes[0, 0]
    panel(ax, "a", "Median AP gain relative to the protected anchor")
    heatmap(ax, ap.to_numpy(float), EVENTS, [f"H{h}" for h in HORIZONS], fmt="+.3f", cbar_label="AP gain")

    ax = axes[0, 1]
    panel(ax, "b", "Additional Q98 exceedance days at a fixed top-2% budget")
    x = np.arange(len(HORIZONS))
    for h_idx, h in enumerate(HORIZONS):
        vals = q98_seed.loc[q98_seed["horizon"].eq(h), "top2_hits_gain"].to_numpy(float)
        jitter = np.linspace(-0.08, 0.08, len(vals))
        ax.scatter(np.full(len(vals), h_idx) + jitter, vals, s=28, color=COLORS["dala"], alpha=0.72, zorder=3)
        ax.plot([h_idx - 0.18, h_idx + 0.18], [np.median(vals)] * 2, color=COLORS["ink"], lw=2.2, zorder=4)
        ax.annotate(
            f"{np.median(vals):,.0f}",
            (h_idx, np.median(vals)),
            xytext=(0, 8),
            textcoords="offset points",
            va="bottom",
            ha="center",
            fontsize=8.5,
            fontweight="bold",
        )
    ax.axhline(0, color=COLORS["ink"], linewidth=0.8)
    ax.set_xticks(x, [f"H{h}" for h in HORIZONS])
    ax.set_ylabel("Extra exceedance days\n(10-source total per DALA seed)")
    clean_axes(ax)

    ax = axes[1, 0]
    panel(ax, "c", "Q98 fixed-budget gain by reporting group")
    heatmap(
        ax,
        source_matrix.to_numpy(float),
        [source.replace("/", "-") for source in SOURCES],
        [f"H{h}" for h in HORIZONS],
        fmt="+.0f",
        cbar_label="Median extra exceedance days",
    )

    ax = axes[1, 1]
    panel(ax, "d", "Lower-tail AP gain and directional consistency")
    annotations = np.empty(lower.shape, dtype=object)
    for r in range(lower.shape[0]):
        for c in range(lower.shape[1]):
            annotations[r, c] = f"{lower.iloc[r, c]:+.4f}\n{int(nn.iloc[r, c])}/50"
    heatmap(
        ax,
        lower.to_numpy(float),
        EVENTS,
        [f"H{h}" for h in HORIZONS],
        cmap="YlGnBu",
        text=annotations,
        cbar_label="10th-percentile AP gain",
    )
    ax.text(0.5, -0.17, "Cell text: 10th-percentile AP gain; nonnegative source-seed units", transform=ax.transAxes, ha="center", fontsize=8)

    save_figure(fig, 2)
    return {"Fig2a_AP_grid": cell, "Fig2b_Q98_seed_hits": q98_seed, "Fig2c_Q98_source_hits": q98_source, "Fig2d_unit_stability": cell}


def build_figure2(primary: pd.DataFrame) -> dict[str, pd.DataFrame]:
    compare = pd.read_csv(DATA / "longlead_dala_vs_current_flow_unit.csv")
    compare_summary = (
        compare.groupby(["event", "horizon"], sort=False)
        .agg(
            units=("source", "size"),
            median_current_gain=("daily_ap_gain_vs_anchor__current_flow_only", "median"),
            median_dala_gain=("daily_ap_gain_vs_anchor__op_lead_adaptive_direct", "median"),
            min_current_gain=("daily_ap_gain_vs_anchor__current_flow_only", "min"),
            min_dala_gain=("daily_ap_gain_vs_anchor__op_lead_adaptive_direct", "min"),
        )
        .reset_index()
    )
    regimes_seed = pd.read_csv(DATA / "mechanism_regime_by_seed_anchor.csv")
    regimes = (
        regimes_seed.groupby(["event", "horizon", "regime"], sort=False)
        .agg(
            seeds=("seed", "nunique"),
            median_new_hits=("new_hits", "median"),
            min_new_hits=("new_hits", "min"),
            max_new_hits=("new_hits", "max"),
            median_new_hit_share=("new_hit_share", "median"),
        )
        .reset_index()
    )
    subset = pd.read_csv(DATA / "mechanism_subset_cell_summary_anchor.csv")
    subset = subset[subset["subset"].eq("near_below")].copy()
    cells = [f"{e}-H{h}" for e in EVENTS for h in (5, 7)]

    fig, axes = plt.subplots(2, 2, figsize=(11.4, 8.4), constrained_layout=True)
    ax = axes[0, 0]
    panel(ax, "a", "Long-lead AP gain: raw state versus conditional admission")
    x = np.arange(len(cells))
    width = 0.36
    lookup = compare_summary.set_index(["event", "horizon"])
    current_vals = [lookup.loc[(e, h), "median_current_gain"] for e in EVENTS for h in (5, 7)]
    dala_vals = [lookup.loc[(e, h), "median_dala_gain"] for e in EVENTS for h in (5, 7)]
    ax.bar(x - width / 2, current_vals, width, color=COLORS["current"], label="Current discharge only", zorder=3)
    ax.bar(x + width / 2, dala_vals, width, color=COLORS["dala"], label="DALA-Hydro", zorder=3)
    ax.axhline(0, color=COLORS["ink"], linewidth=0.9)
    ax.set_xticks(x, cells, rotation=25, ha="right")
    ax.set_ylabel("Median AP gain vs anchor")
    ax.legend(frameon=False, fontsize=8.5, loc="upper right")
    clean_axes(ax)

    ax = axes[0, 1]
    panel(ax, "b", "Where newly detected long-lead exceedance days originate")
    regime_order = ["far_below", "near_below", "above"]
    labels = {"far_below": "Far below", "near_below": "Near below", "above": "Already above"}
    colors = {"far_below": COLORS["far"], "near_below": COLORS["near"], "above": COLORS["above"]}
    for idx, regime in enumerate(regime_order):
        vals = []
        for e in EVENTS:
            for h in (5, 7):
                row = regimes[(regimes["event"].eq(e)) & (regimes["horizon"].eq(h)) & (regimes["regime"].eq(regime))]
                vals.append(float(row["median_new_hits"].iloc[0]) if len(row) else 0.0)
        ax.bar(x + (idx - 1) * 0.25, vals, 0.24, color=colors[regime], label=labels[regime], zorder=3)
    ax.set_xticks(x, cells, rotation=25, ha="right")
    ax.set_ylabel("New true detections\n(median 10-source total per seed)")
    ax.legend(frameon=False, fontsize=8.2, ncol=3, loc="upper right")
    clean_axes(ax)

    ax = axes[1, 0]
    panel(ax, "c", "Near-below contribution to new long-lead detections")
    share = regimes[regimes["regime"].eq("near_below")].pivot(index="event", columns="horizon", values="median_new_hit_share").reindex(index=EVENTS, columns=[5, 7])
    heatmap(ax, 100 * share.to_numpy(float), EVENTS, ["H5", "H7"], fmt=".1f", cmap="YlOrBr", cbar_label="Share of new detections (%)")

    ax = axes[1, 1]
    panel(ax, "d", "Ranking gain within the near-below subset")
    near_ap = subset.pivot(index="event", columns="horizon", values="median_ap_gain_vs_anchor").reindex(index=EVENTS, columns=[5, 7])
    heatmap(ax, near_ap.to_numpy(float), EVENTS, ["H5", "H7"], fmt="+.3f", cbar_label="Median subset AP gain")

    save_figure(fig, 3)
    return {"Fig3a_current_vs_DALA": compare, "Fig3bc_regimes_seed": regimes_seed, "Fig3bc_regimes_summary": regimes, "Fig3d_near_below_AP": subset}


def build_figure3(primary: pd.DataFrame) -> dict[str, pd.DataFrame]:
    nested = pd.read_csv(DATA / "nested_loso_by_fold.csv")
    nested_cells = pd.read_csv(DATA / "nested_loso_cell_summary.csv")
    post = pd.read_csv(DATA / "postlock_lamahice_anchor_unit_metrics.csv")
    post_cells = pd.read_csv(DATA / "postlock_lamahice_anchor_summary.csv")
    boot = pd.read_csv(DATA / "source_block_bootstrap_anchor.csv")
    seed = source_seed_rollup(primary)

    nmat = nested_cells.pivot(index="event", columns="horizon", values="median_heldout_daily_ap_gain").reindex(index=EVENTS, columns=HORIZONS)
    pmat = post_cells.pivot(index="event", columns="horizon", values="median_ap_gain_vs_anchor").reindex(index=EVENTS, columns=HORIZONS)
    emat = seed.groupby(["event", "horizon"], sort=False)["episode_gain"].median().unstack("horizon").reindex(index=EVENTS, columns=HORIZONS)

    fig, axes = plt.subplots(2, 2, figsize=(11.4, 8.4), constrained_layout=True)
    ax = axes[0, 0]
    panel(ax, "a", "Nested source-wise candidate selection")
    heatmap(ax, nmat.to_numpy(float), EVENTS, [f"H{h}" for h in HORIZONS], fmt="+.3f", cbar_label="Held-out median AP gain")

    ax = axes[0, 1]
    panel(ax, "b", "Post-lock LamaH-Ice evaluation")
    heatmap(ax, pmat.to_numpy(float), EVENTS, [f"H{h}" for h in HORIZONS], fmt="+.3f", cbar_label="Median AP gain")

    ax = axes[1, 0]
    panel(ax, "c", "Archive-block bootstrap uncertainty")
    order = [(e, h) for e in EVENTS for h in HORIZONS]
    b = boot.set_index(["event", "horizon"]).loc[order].reset_index()
    x = np.arange(len(b))
    y = b["observed_median_ap_gain_vs_anchor"].to_numpy(float)
    lo = b["ci025"].to_numpy(float)
    hi = b["ci975"].to_numpy(float)
    ax.errorbar(x, y, yerr=np.vstack([y - lo, hi - y]), fmt="o", color=COLORS["dala"], ecolor=COLORS["ink"], capsize=3, markersize=5, zorder=3)
    ax.axhline(0, color=COLORS["ink"], linewidth=0.9)
    ax.set_xticks(x, [f"{e}\nH{h}" for e, h in order])
    ax.set_ylabel("AP gain (95% block-bootstrap interval)")
    clean_axes(ax)

    ax = axes[1, 1]
    panel(ax, "d", "Additional unique high-flow episodes at equal alert volume")
    heatmap(ax, emat.to_numpy(float), EVENTS, [f"H{h}" for h in HORIZONS], fmt="+.0f", cbar_label="Median extra episodes per DALA seed")

    save_figure(fig, 4)
    return {"Fig4a_nested_folds": nested, "Fig4a_nested_summary": nested_cells, "Fig4b_postlock_units": post, "Fig4b_postlock_summary": post_cells, "Fig4c_block_bootstrap": boot, "Fig4d_episode_seed_totals": seed}


def safe_sheet_name(name: str) -> str:
    return name[:31]


def write_source_data(frames: dict[str, pd.DataFrame]) -> Path:
    path = OUT / "Result_Figure_Source_Data.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    index = wb.create_sheet("Index")
    index.append(["Sheet", "Contents"])
    descriptions = {
        "Fig2a_AP_grid": "Figure 2a and 1d: primary cell AP gain and stability summaries.",
        "Fig2b_Q98_seed_hits": "Figure 2b: all five 10-source Q98 fixed-budget totals.",
        "Fig2c_Q98_source_hits": "Figure 2c: Q98 source-level fixed-budget summaries.",
        "Fig2d_unit_stability": "Figure 2d: minimum and nonnegative AP units per cell.",
        "Fig3a_current_vs_DALA": "Figure 3a: all source-seed long-lead comparisons.",
        "Fig3bc_regimes_seed": "Figure 3b-c: seed-level regime counts and shares.",
        "Fig3bc_regimes_summary": "Figure 3b-c: plotted regime summaries.",
        "Fig3d_near_below_AP": "Figure 3d: near-below subset AP summaries.",
        "Fig4a_nested_folds": "Figure 4a: all nested held-out source folds.",
        "Fig4a_nested_summary": "Figure 4a: plotted nested cell summaries.",
        "Fig4b_postlock_units": "Figure 4b: all post-lock LamaH-Ice seed units.",
        "Fig4b_postlock_summary": "Figure 4b: plotted post-lock summaries.",
        "Fig4c_block_bootstrap": "Figure 4c: archive-block bootstrap point estimates and intervals.",
        "Fig4d_episode_seed_totals": "Figure 4d: all five seed-level episode totals.",
    }
    for key, frame in frames.items():
        frame = frame.copy()
        for column in frame.select_dtypes(include="object").columns:
            frame[column] = frame[column].replace({"NZ/FI": "NZ-FI"})
        sheet_name = safe_sheet_name(key)
        index.append([sheet_name, descriptions.get(key, key)])
        ws = wb.create_sheet(sheet_name)
        ws.append(list(frame.columns))
        for row in frame.itertuples(index=False, name=None):
            ws.append([None if (isinstance(v, float) and (math.isnan(v) or math.isinf(v))) else v for v in row])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="006D67")
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for col_idx, column in enumerate(ws.columns, 1):
            max_len = min(42, max(len(str(cell.value)) if cell.value is not None else 0 for cell in column) + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, max_len)
    for cell in index[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="006D67")
    index.column_dimensions["A"].width = 32
    index.column_dimensions["B"].width = 95
    index.freeze_panes = "A2"
    wb.save(path)
    return path


def build() -> Path:
    setup_style()
    primary = load_primary()
    frames: dict[str, pd.DataFrame] = {}
    frames.update(build_figure1(primary))
    frames.update(build_figure2(primary))
    frames.update(build_figure3(primary))
    write_source_data(frames)
    return OUT


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build Journal of Hydrology result Figures 2–4.")
    parser.add_argument("--data-dir", type=Path, default=DATA, help="Directory containing the derived CSV tables.")
    parser.add_argument("--out-dir", type=Path, default=OUT, help="Directory for the figures and Source_Data.xlsx.")
    args = parser.parse_args()
    DATA = args.data_dir.resolve()
    OUT = args.out_dir.resolve()
    print(build())
